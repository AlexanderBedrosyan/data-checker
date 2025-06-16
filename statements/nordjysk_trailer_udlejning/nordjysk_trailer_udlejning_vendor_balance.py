import os
import sys
from openpyxl import load_workbook
sys.path.append(os.path.abspath(os.path.join(os.path.dirname(__file__), '..')))
from statements.basic_model import *


class NordjyskTrailerModel(BasicModel):

    def vendor_balance(self):
        folder_path = self.folder_path
        bc_balance_path = self.find_pdf_file()

        workbook = load_workbook(filename=bc_balance_path)

        sheet = workbook.active
        balance = {}

        for row in sheet.iter_rows(values_only=True):
            current_row = row[0]
            current_row = current_row.split(' ')
            if len(current_row) < 2:
                continue

            if self.regex_finder(current_row[0]) and current_row[1] in ['Faktura:', 'Credit Memo:', 'Kreditnota:']:

                invoice = current_row[2]
                amount = None
                if len(current_row) > 5:
                    amount = current_row[-2]
                else:
                    amount = current_row[-1]
                amount = amount.replace('.', '').replace(',', '.')
                balance[invoice] = amount

        return balance
    

nordjysk_trailer = NordjyskTrailerModel()