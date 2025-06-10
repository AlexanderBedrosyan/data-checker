import os
import sys
from openpyxl import load_workbook
from statements.basic_model import BasicModel
sys.path.append(os.path.abspath(os.path.join(os.path.dirname(__file__), '..')))


class AhFragt(BasicModel):

    def vendor_balance(self):
        folder_path = self.folder_path
        bc_balance_path = self.find_pdf_file()

        workbook = load_workbook(filename=bc_balance_path)

        sheet = workbook.active
        balance = {}

        for row in sheet.iter_rows(values_only=True):
            current_row = row[0].split(' ')
            try:
                if current_row[0] == 'Kontoudtog':
                    continue
                if self.regex_finder(current_row[2]):
                    doc_num = str(current_row[0])
                    amount = current_row[-2].replace('.', '').replace(',', '.')
                    balance[doc_num] = amount
            except:
                continue

        return balance


ah_fragt = AhFragt()
