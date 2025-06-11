import os
import sys
from openpyxl import load_workbook
sys.path.append(os.path.abspath(os.path.join(os.path.dirname(__file__), '..')))
from statements.basic_model import *


class EgsjModel(BasicModel):
            
    def vendor_balance(self):
        folder_path = self.folder_path
        bc_balance_path = self.find_pdf_file()

        workbook = load_workbook(filename=bc_balance_path)

        sheet = workbook.active
        balance = {}

        for row in sheet.iter_rows(values_only=True):
            curr_row = row[0].split(' ')      

            if self.regex_finder(curr_row[0]) and self.regex_finder(curr_row[4]) and len(curr_row) >= 7:
                invoice_num = curr_row[1]
                amount = None
                if len(curr_row) == 7:
                    amount = curr_row[-2].replace('.', '')
                else:
                    amount = curr_row[-3].replace('.', '')
                amount = amount.replace(',', '.')
                balance[invoice_num] = amount

        return balance
    

egsj = EgsjModel()