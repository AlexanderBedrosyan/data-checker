import pdfplumber
import pandas as pd
import os
from typing import List
from openpyxl import load_workbook, Workbook


def find_pdf_file(current_path):
    """
    Find all PDF files in the current folder.
    Add them in list and return the list
    """
    all_pdf_files = []
    for file in os.listdir(current_path):
        if file.lower().endswith('.pdf'):
            full_path = os.path.join(current_path, file)
            all_pdf_files.append(full_path)
    return all_pdf_files


def append_df_to_excel(df, excel_path):
    """
    Convert the details from the PDF to excel, which are already converted in df (first argument in the fuction).
    Save them to the excel_path (second argument in the function)
    """
    if os.path.exists(excel_path):  
        book = load_workbook(excel_path)
        sheet = book.active  
        start_row = sheet.max_row
    else:
        book = Workbook()
        sheet = book.active
        sheet.title = 'Sheet1'
        start_row = 0

    for index, row in df.iterrows():
        sheet.append(row.tolist())

    book.save(excel_path)
        

def pdf_to_excel(pdf_paths=List, excel_path=str):
    """
    Read, order and prepare the pdf files for converting to excel.
    Using find_pdf_file function to check all pdf include in the current folder.
    """
    all_pages_text = []

    for pdf_path in pdf_paths:
        with pdfplumber.open(pdf_path) as pdf:
            for page in pdf.pages:
                text = ''
                text = page.extract_text()
                all_pages_text.append(text)
        
    full_text = "\n".join(all_pages_text)

    rows = [line.split('\t') for line in full_text.split('\n') if line.strip() != '']

    pd.DataFrame().to_excel(excel_path)

    for row in rows:
        try:
            df = pd.DataFrame(row)
            append_df_to_excel(df, excel_path)
        except Exception as e:
            print(f"Error processing row: {e}")
            continue


current_folder_path = os.getcwd().split("\\")
folder_path = '\\'.join(current_folder_path[:-1]) + "\\uploads"
pdf_paths = find_pdf_file(folder_path)
excel_path = folder_path + '\\' + 'received_vendor_balance.xlsx'
pdf_to_excel(pdf_paths, excel_path)
