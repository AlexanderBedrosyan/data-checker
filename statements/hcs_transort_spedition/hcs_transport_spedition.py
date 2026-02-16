import os
import sys
from openpyxl import load_workbook
sys.path.append(os.path.abspath(os.path.join(os.path.dirname(__file__), '..')))
from basic_model import *


class Hcs_Transport_Spedition(BasicModel):

    def vendor_balance(self):
        folder_path = self.folder_path
        bc_balance_path = self.find_pdf_file()

        workbook = load_workbook(filename=bc_balance_path)
        sheet = workbook.active
        is_started = False
        balance = {}

        for row in sheet.iter_rows(values_only=True):
            current_row = row[0].split(' ')
            if self.regex_finder(current_row[0]):
                inv_num = ''
                if not current_row[1].startswith('7'):
                    inv_num = current_row[1]
                else:
                    inv_num = current_row[-2]
                amount = current_row[-1].replace('.', '')
                amount = amount.replace(',', '.')
                balance[inv_num] = amount

        return balance
    

hcs_transport_spedition = Hcs_Transport_Spedition()
