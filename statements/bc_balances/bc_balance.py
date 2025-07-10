import os
from openpyxl import load_workbook


CURRENT_DIR = os.path.dirname(os.path.abspath(__file__))


def find_pdf_file(current_path):
    all_files = []
    for file in os.listdir(current_path):
        if file.endswith('.xlsx') and (file.lower().startswith('vendor') or file.lower().startswith('customer')):
            all_files.append(current_path + '/' + file)
    return all_files
        

def position_finder(curr_workbook, type_of_balance=str):
    pos_inv_number = None
    pos_amount = None
    pos_doc_type = None

    for i in range(len(curr_workbook[1])):
        document_mapper = {
            'customer': 'Document No.',
            'vendor': 'External Document No.'
        }
        if curr_workbook[1][i].value == document_mapper[type_of_balance]:
            pos_inv_number = i
        elif curr_workbook[1][i].value == 'Original Amount':
            pos_amount = i
        elif curr_workbook[1][i].value == 'Document Type':
            pos_doc_type = i
    return pos_inv_number, pos_amount, pos_doc_type


def bc_balance():
    NEEDED_DOCUMENTS = ['Invoice', 'Credit Memo']
    folder_path = os.path.join(os.path.dirname(CURRENT_DIR).rstrip("statements"), 'uploads')
    bc_balances_path = find_pdf_file(folder_path)
    balance = {}

    for path in bc_balances_path:
        workbook = load_workbook(filename=path)

        sheet = workbook.active
        type_of_balance = 'customer' if 'customer' in path.lower() else 'vendor'

        pos_inv_number, pos_amount, pos_doc_type = position_finder(sheet, type_of_balance)

        for row in sheet.iter_rows(values_only=True):
            document_type = row[pos_doc_type]

            if document_type not in NEEDED_DOCUMENTS:
                continue

            doc_number = row[pos_inv_number]
            # ---------------------------
            if doc_number.startswith('400') and document_type == 'Invoice':
                doc_number = row[-4]
            # ---------------------------
            if doc_number not in balance:
                balance[doc_number] = 0
            amount = row[pos_amount]
            balance[doc_number] += float(amount)

    return balance


def find_account_number():
    folder_path = os.path.join(os.path.dirname(CURRENT_DIR).rstrip("statements"), 'uploads')
    bc_balances_path = find_pdf_file(folder_path)
    account = []
    for path in bc_balances_path:
        workbook = load_workbook(filename=path)

        sheet = workbook.active

        index_placed = 0

        for i in range(len(sheet[1])):
            if sheet[1][i].value in ("Vendor No.", "Customer No."):
                index_placed = i
                break
        
        for row in sheet.iter_rows(values_only=True):
            if row[index_placed] not in ("Vendor No.", "Customer No.") and row[index_placed] not in account:
                account.append(row[index_placed])

    return account

bc_balance()