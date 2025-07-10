import os
import sys
from openpyxl import load_workbook
sys.path.append(os.path.abspath(os.path.join(os.path.dirname(__file__), '..')))
from statements.basic_model import *


class SvendSorensen(BasicModel):

    def vendor_balance(self):
        folder_path = self.folder_path
        bc_balance_path = self.find_pdf_file()

        workbook = load_workbook(filename=bc_balance_path)

        sheet = workbook['Sheet1']
        balance = {}

        for row in sheet.iter_rows(values_only=True):
            curr_row = row[0].split(' ')

            if not self.regex_finder(curr_row[0]) or curr_row[1] not in ["Faktura:", "Kreditnota:"]:
                continue

            invoice_num = curr_row[2]
            amount = None
            try:
                amount = row[-3][3:].replace('.', '')
            except:
                amount = float(curr_row[-2].replace('.', '').replace(',', '.'))

            if invoice_num not in balance:
                balance[invoice_num] = 0

            try:
                balance[invoice_num] += float(amount)
            except ValueError:
                print("Error in row 35 of Svend E Sorensen vendor Balance.")

        return balance
    

svend_e_sorensen = SvendSorensen()