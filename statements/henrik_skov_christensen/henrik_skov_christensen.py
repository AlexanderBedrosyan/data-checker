import os
import sys
from openpyxl import load_workbook

sys.path.append(os.path.abspath(os.path.join(os.path.dirname(__file__), '..')))
from statements.basic_model import *


class HenrikSkovChristensen(BasicModel):

    def vendor_balance(self):
        folder_path = self.folder_path
        bc_balance_path = self.find_pdf_file()
        workbook = load_workbook(filename=bc_balance_path)

        sheet = workbook.active
        balance = {}

        for row in sheet.iter_rows(values_only=True):
            row = row[0].split(' ')
            try:
                if not self.regex_finder(row[0]) and len(row) != 7:
                    continue
                doc_num = str(row[1])
                amount = row[-2].replace('.', '').replace(',', '.')
                if '-' in amount:
                    amount = amount.replace('-', '')
                    amount = '-' + amount
                balance[doc_num] = amount
            except:
                continue
        return balance


henrik_skov_christensen = HenrikSkovChristensen()