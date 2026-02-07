from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from werkzeug.datastructures import FileStorage
import re


def extract_sheets_to_dicts(
    uploaded_file: FileStorage,
    sheet_names=("ICTemplateDataBS", "ReceivablePayable"),
):
    """
    Връща dict със sheet_name -> dict
    key = първа колона
    value = list с останалите колони (включително празни)
    """

    wb = load_workbook(uploaded_file, data_only=True)

    result = {}

    for sheet_name in sheet_names:
        if sheet_name not in wb.sheetnames:
            raise ValueError(f"Sheet '{sheet_name}' не съществува")

        ws = wb[sheet_name]
        sheet_dict = {}

        for row in ws.iter_rows(values_only=True):
            if not row:
                continue

            key = row[0]

            # ако първата колона е празна → пропускаме реда
            if key is None:
                continue

            if key not in sheet_dict:
                sheet_dict[key] = []

            if sheet_name == "ICTemplateDataBS":
                # за ICTemplateDataBS искаме само последната колона (ако има повече от 2)
                value = str(row[1]) + ' ' + str(row[-1])
                sheet_dict[key].append(value)
                continue

            values = [
                "" if cell is None else cell
                for cell in row[1:]
            ]

            sheet_dict[key] = values

        result[sheet_name] = sheet_dict
    return result


def column4_finder(first_row):
    for index, header in enumerate(first_row):
        if header.lower() == "column4" in str(header).lower():
            return index


def find_last_difference_column(header_row):
    """
    Find the last column that starts with 'Difference' in the header row.
    Returns the column index and the next difference number.
    """
    last_diff_index = -1
    max_diff_number = 0
    
    for index, header in enumerate(header_row):
        if header and isinstance(header, str) and header.startswith('Difference'):
            last_diff_index = index
            # Extract number from DifferenceN
            match = re.search(r'Difference(\d+)', header)
            if match:
                diff_number = int(match.group(1))
                if diff_number > max_diff_number:
                    max_diff_number = diff_number
    
    next_diff_number = max_diff_number + 1
    return last_diff_index, next_diff_number


def ic_data_modifier(ic_data):
    for company_name in list(ic_data.keys())[::-1]:
        first_code, second_code = company_name.split(" - ")

        test_code = f"{second_code} - {first_code}"
        if test_code in ic_data:
            ic_data.pop(company_name)
    return ic_data

def ic_template_data_bs_preparation(data):
    """
    Prepare ICTemplateDataBS data.
    Returns a dict with company names as keys and their values.
    """
    final_report = {}

    for comp_name, details in data.items():
        if comp_name == "Company":  # Skip header row
            continue
            
        for element in details:
            element_parts = element.split()
            if len(element_parts) < 2:
                continue
                
            new_name = comp_name + " - " + element_parts[0]
            check_name = element_parts[0] + " - " + comp_name

            if check_name in final_report:
                continue

            try:
                final_report[new_name] = float(element_parts[-1])
            except (ValueError, IndexError):
                final_report[new_name] = 0.0

    return final_report


def process_sheet_pair(wb, ic_template_sheet_name, target_sheet_name, yellow_fill, green_fill):
    """
    Process one IC template sheet with its corresponding target sheet.
    
    Args:
        wb: The workbook object
        ic_template_sheet_name: Name of the IC template sheet (e.g., "ICTemplateDataBS")
        target_sheet_name: Name of the target sheet (e.g., "ReceivablePayable")
        yellow_fill: PatternFill object for yellow color
        green_fill: PatternFill object for green color
    """
    # Process target sheet
    if target_sheet_name not in wb.sheetnames:
        raise ValueError(f"Sheet '{target_sheet_name}' not found")
    
    target_sheet = wb[target_sheet_name]
    
    # Remove all tables in the sheet to prevent corruption when inserting rows/columns
    if target_sheet.tables:
        tables_to_remove = list(target_sheet.tables.keys())
        for table_name in tables_to_remove:
            del target_sheet.tables[table_name]
    target_sheet = wb[target_sheet_name]
    
    # Remove all tables in the sheet to prevent corruption when inserting rows/columns
    if target_sheet.tables:
        tables_to_remove = list(target_sheet.tables.keys())
        for table_name in tables_to_remove:
            del target_sheet.tables[table_name]
    
    # Get header row (first row)
    header_row = []
    company_col_index = None
    column4_col_index = None
    
    for col_idx, cell in enumerate(target_sheet[1], start=1):
        header_value = cell.value
        header_row.append(header_value)
        
        if header_value == "Company":
            company_col_index = col_idx
        elif header_value == "Column4":
            column4_col_index = col_idx
    
    # Find last Difference column
    last_diff_col_index, next_diff_number = find_last_difference_column(header_row)
    
    if last_diff_col_index == -1:
        raise ValueError(f"No Difference column found in the header of {target_sheet_name}")
    
    # INSERT new column after the last Difference column
    # last_diff_col_index is 0-based, so +1 for column number, +1 for next position
    new_diff_col_index = last_diff_col_index + 2
    target_sheet.insert_cols(new_diff_col_index)
    
    # Add new Difference column header
    new_diff_header = f"Difference{next_diff_number}"
    target_sheet.cell(row=1, column=new_diff_col_index, value=new_diff_header)
    
    # Update column4_col_index if it was after the insertion point
    if column4_col_index and column4_col_index >= new_diff_col_index:
        column4_col_index += 1
    
    # Process IC template sheet
    if ic_template_sheet_name not in wb.sheetnames:
        raise ValueError(f"Sheet '{ic_template_sheet_name}' not found")
    
    # Extract IC template data
    ic_data = {}
    ic_sheet = wb[ic_template_sheet_name]
    
    # Remove all tables in IC template sheet as well
    if ic_sheet.tables:
        tables_to_remove = list(ic_sheet.tables.keys())
        for table_name in tables_to_remove:
            del ic_sheet.tables[table_name]
    # Remove all tables in IC template sheet as well
    if ic_sheet.tables:
        tables_to_remove = list(ic_sheet.tables.keys())
        for table_name in tables_to_remove:
            del ic_sheet.tables[table_name]
    
    for row in ic_sheet.iter_rows(min_row=2, values_only=True):  # Skip header
        if not row or row[0] is None:
            continue
        
        company_name = f"{row[0]} - {row[1]}"

        # Assuming format: Company | SomeCode | Value
        if len(row) >= 3:
            code = str(row[1]) if row[1] else ""
            value = row[-1]  # Last column value
            
            try:
                value_float = float(value) if value else 0.0
                # Only add if value is not zero
                ic_data[company_name] = round(value_float, 2)
            except (ValueError, TypeError):
                pass
    
    ic_data = ic_data_modifier(ic_data)
    # Track which companies from IC data were found in target sheet
    found_companies = set()

    def find_insert_row(sheet, company_col, company_name, start_row=2):
        """Find the correct row position to insert a company name alphabetically"""
        target_key = str(company_name).strip().casefold()
        max_row = sheet.max_row

        for row_idx in range(start_row, max_row + 1):
            existing = sheet.cell(row=row_idx, column=company_col).value
            if existing is None or str(existing).strip() == "":
                continue

            existing_key = str(existing).strip().casefold()
            if existing_key > target_key:
                return row_idx

        return max_row + 1
    
    # Process each row in target sheet (starting from row 2)
    for row_idx in range(2, target_sheet.max_row + 1):
        company_cell = target_sheet.cell(row=row_idx, column=company_col_index)
        company_name = company_cell.value
        
        # Skip rows with empty company names
        if not company_name or str(company_name).strip() == "":
            continue
        
        # Get value from last Difference column (before insertion)
        last_diff_cell = target_sheet.cell(row=row_idx, column=last_diff_col_index + 1)
        last_diff_value = last_diff_cell.value
        
        try:
            last_diff_value = float(last_diff_value) if last_diff_value else 0.0
        except (ValueError, TypeError):
            last_diff_value = 0.0
        
        # Look for matching company in IC data
        matching_ic_value = None
        for ic_company, ic_value in ic_data.items():
            # Skip if ic_company is None or empty
            if not ic_company:
                continue
            if company_name == ic_company or company_name in ic_company or ic_company in company_name:
                matching_ic_value = ic_value
                found_companies.add(ic_company)
                break
        
        # If no match found in IC data, skip this row
        if matching_ic_value is None:
            continue
        
        # Get Column4 value
        column4_cell = target_sheet.cell(row=row_idx, column=column4_col_index) if column4_col_index else None
        column4_value = column4_cell.value if column4_cell else None
        
        # Apply logic
        new_cell = target_sheet.cell(row=row_idx, column=new_diff_col_index)
        
        if abs(matching_ic_value) > 50000 and column4_value is None and target_sheet_name != "FTEs":  # Different values (with small tolerance)
            # Values are different -> add value and color yellow
            new_cell.value = matching_ic_value
            new_cell.fill = yellow_fill
        elif abs(matching_ic_value) > 0 and column4_value is None and target_sheet_name == "FTEs":
            # Values are different -> add value and color yellow
            new_cell.value = matching_ic_value
            new_cell.fill = yellow_fill
        else:
            new_cell.value = matching_ic_value
            new_cell.fill = green_fill
    
    # Add missing companies from IC data as new rows (alphabetically)
    missing_companies = set(ic_data.keys()) - found_companies

    if missing_companies:
        # Insert missing companies in alphabetical order
        for company_name in sorted(missing_companies, key=lambda x: str(x).strip().casefold()):
            ic_value = ic_data[company_name]

            if abs(ic_value) < 50000 and target_sheet_name != "FTEs":
                continue  # Skip companies with low value

            if abs(ic_value) == 0 and target_sheet_name == "FTEs":
                continue  # Skip companies with low value in FTEs sheet
            
            # Find correct position to insert based on alphabetical order
            insert_row = find_insert_row(target_sheet, company_col_index, company_name)
            
            # Insert a new row at the correct position
            target_sheet.insert_rows(insert_row)
            
            # Add company name in Company column
            target_sheet.cell(row=insert_row, column=company_col_index, value=company_name)
            
            # Add value in the new Difference column with yellow color (since it's new)
            new_cell = target_sheet.cell(row=insert_row, column=new_diff_col_index, value=ic_value)
            new_cell.fill = yellow_fill
    
    # Clear data validation rules to prevent corruption
    if hasattr(target_sheet, 'data_validations') and target_sheet.data_validations:
        target_sheet.data_validations = None
    
    if hasattr(ic_sheet, 'data_validations') and ic_sheet.data_validations:
        ic_sheet.data_validations = None


def process_excel_with_difference_logic(uploaded_file: FileStorage, output_path: str):
    """
    Main function to process the Excel file with the Difference column logic for all sheet pairs.
    
    Processes the following sheet pairs:
    - ICTemplateDataBS → ReceivablePayable
    - ICTemplateDataICL → Loan
    - ICTemplateDataCP → Cash Pool
    - ICTemplateDataFTE → FTEs
    
    Steps:
    1. Load the Excel file with formatting
    2. For each sheet pair:
       a. Find the last Difference column
       b. INSERT a new Difference column (not overwrite)
       c. Compare values from IC template and populate the new column
       d. Apply color formatting (yellow or green)
       e. Add missing companies as new rows
    3. Save the modified file
    """
    # Load workbook with formatting preserved
    wb = load_workbook(uploaded_file)
    
    # Define color fills
    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    green_fill = PatternFill(start_color="FF00B050", end_color="FF00B050", fill_type="solid")
    
    # Define sheet pairs: IC template sheet → Target sheet
    sheet_pairs = [
        ("ICTemplateDataBS", "ReceivablePayable"),
        ("ICTemplateDataICL", "Loan"),
        ("ICTemplateDataCP", "Cash Pool"),
        ("ICTemplateDataFTE", "FTEs")
    ]
    
    # Process each sheet pair
    for ic_template_sheet, target_sheet in sheet_pairs:
        try:
            process_sheet_pair(wb, ic_template_sheet, target_sheet, yellow_fill, green_fill)
        except ValueError as e:
            # Log the error but continue with other sheets
            print(f"Warning: Could not process {ic_template_sheet} → {target_sheet}: {e}")
            continue
    
    # Save the modified workbook
    wb.save(output_path)
    return output_path